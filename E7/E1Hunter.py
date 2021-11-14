from pyhunter import PyHunter
from openpyxl import Workbook
import getpass

# Participantes:
# 1.- Alejandro Cavazos Valdés
# 2.- Julio Alonso Grimaldo Mercado
# 3.- Ángel Ivan Celaya Garcia
# 4.- Julio Gerardo Cazares Gonzalez


def Busqueda(organizacion):
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 50, cuidado!
    resultado = hunter.domain_search(
        company=organizacion, limit=1, emails_type="personal"
    )
    return resultado


def GuardarInformacion(datosEncontrados, organizacion):
    # libro = openxyl.Workbook()
    libro = Workbook()
    pagina = libro.active
    pagina1 = libro.create_sheet(organizacion)
    if len(libro.sheetnames) > 1:
        libro.active = 1
        hoja = libro.active
    else:
        hoja = libro.active
    print("La página activa es:", hoja.title)
    datosImportantes = ("domain", "organization", "country", "emails")
    datosImportantes2 = (
        "value", "type", "sources", "first_name", "last_name", "phone_number"
    )
    datosImportantes3 = ("domain", "uri", "extracted_on", "last_seen_on")
    data1 = list()
    data2 = list()
    data3 = list()
    hoja["A1"] = "domain"
    for i in range(1, 4):
        hoja[f"A{1+i}"] = datosImportantes[i]
    for val in range(len(datosImportantes2)):
        hoja[f"A{4+val}"] = datosImportantes2[val]
    hoja["A6"] = "first_name"
    hoja["A7"] = "last_name"
    hoja["A8"] = "phone_number"
    for val2 in range(len(datosImportantes3)):
        hoja[f"A{8+val2}"] = datosImportantes3[val2]
    for y in datosImportantes:
        x = datosEncontrados[y]
        data1.append(x)
        value = type(x).__name__
        if value == "list":
            newdata = datosEncontrados[y]
            for data in datosImportantes2:
                for elem in newdata:
                    i = elem[data]
                    value2 = type(i).__name__
                    data2.append(i)
                    if value2 == "list":
                        lista = elem[data]
                        for elem2 in datosImportantes3:
                            for f in lista:
                                z = f[elem2]
                                data3.append(z)
    data1.pop()
    a = 1
    for i in data1:
        hoja[f"B{a}"] = i
        a += 1
    data2.pop(2)
    a = 4
    for i in data2:
        hoja[f"B{a}"] = i
        a += 1
    a = 8
    for i in data3:
        hoja[f"B{a}"] = i
        a += 1
    libro.save("Hunter" + organizacion + ".xlsx")


print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)
if datosEncontrados is None:
    exit()
else:
    print(datosEncontrados)
    print(type(datosEncontrados))
    GuardarInformacion(datosEncontrados, orga)
